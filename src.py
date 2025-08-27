"""
MSG to Excel Processor
Processes Microsoft Outlook MSG files and automatically pastes extracted data into Excel files.
"""

import os
import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from datetime import datetime

try:
    import extract_msg
except ImportError:
    print("extract_msg not installed. Run: pip install extract-msg")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")

try:
    import pandas as pd
except ImportError:
    print("pandas not installed. Run: pip install pandas")


@dataclass
class EmailData:
    """Data structure for extracted email information"""
    subject: str
    sender: str
    recipient: str
    date: str
    body: str
    attachments: List[str]
    message_id: str


class MSGProcessor:
    """Main class for processing MSG files"""
    
    def __init__(self, log_level: str = "INFO"):
        self.setup_logging(log_level)
        self.logger = logging.getLogger(__name__)
    
    def setup_logging(self, level: str) -> None:
        """Configure logging"""
        logging.basicConfig(
            level=getattr(logging, level.upper()),
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('msg_processor.log'),
                logging.StreamHandler()
            ]
        )
    
    def extract_msg_data(self, msg_file_path: str) -> Optional[EmailData]:
        """Extract data from a single MSG file"""
        try:
            msg = extract_msg.Message(msg_file_path)
            
            # Extract attachments info
            attachments = []
            if hasattr(msg, 'attachments') and msg.attachments:
                for attachment in msg.attachments:
                    if hasattr(attachment, 'longFilename'):
                        attachments.append(attachment.longFilename)
                    elif hasattr(attachment, 'shortFilename'):
                        attachments.append(attachment.shortFilename)
            
            email_data = EmailData(
                subject=msg.subject or "",
                sender=msg.sender or "",
                recipient=msg.to or "",
                date=msg.date.strftime("%Y-%m-%d %H:%M:%S") if msg.date else "",
                body=msg.body or "",
                attachments=attachments,
                message_id=msg.messageId or ""
            )
            
            self.logger.info(f"Successfully extracted data from {msg_file_path}")
            return email_data
            
        except Exception as e:
            self.logger.error(f"Error processing {msg_file_path}: {str(e)}")
            return None
    
    def process_msg_folder(self, folder_path: str) -> List[EmailData]:
        """Process all MSG files in a folder or single MSG file"""
        emails = []
        path = Path(folder_path)
        
        if not path.exists():
            self.logger.error(f"Path does not exist: {folder_path}")
            return emails
        
        # Check if it's a single MSG file
        if path.is_file() and path.suffix.lower() == '.msg':
            self.logger.info(f"Processing single MSG file: {folder_path}")
            email_data = self.extract_msg_data(str(path))
            if email_data:
                emails.append(email_data)
        # If it's a folder, process all MSG files in it
        elif path.is_dir():
            msg_files = list(path.glob("*.msg"))
            self.logger.info(f"Found {len(msg_files)} MSG files in {folder_path}")
            
            for msg_file in msg_files:
                email_data = self.extract_msg_data(str(msg_file))
                if email_data:
                    emails.append(email_data)
        else:
            self.logger.error(f"Path is neither a folder nor an MSG file: {folder_path}")
        
        return emails
    
    def convert_to_json(self, emails: List[EmailData], output_path: str = "emails.json") -> None:
        """Convert extracted email data to JSON"""
        try:
            emails_dict = [
                {
                    "subject": email.subject,
                    "sender": email.sender,
                    "recipient": email.recipient,
                    "date": email.date,
                    "body": email.body,
                    "attachments": email.attachments,
                    "message_id": email.message_id
                }
                for email in emails
            ]
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(emails_dict, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"JSON data saved to {output_path}")
            
        except Exception as e:
            self.logger.error(f"Error saving JSON: {str(e)}")


class ExcelWriter:
    """Class for writing data to Excel files"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def create_excel_from_emails(self, emails: List[EmailData], output_path: str = "emails.xlsx") -> None:
        """Create Excel file from email data"""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Email Data"
            
            # Headers
            headers = ["Subject", "Sender", "Recipient", "Date", "Body Preview", "Attachments", "Message ID"]
            worksheet.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            for email in emails:
                
                body_preview = email.body.replace('\n', ' ').replace('\r', ' ')
                
                attachments_str = ", ".join(email.attachments) if email.attachments else "None"
                
                worksheet.append([
                    email.subject,
                    email.sender,
                    email.recipient,
                    email.date,
                    body_preview,
                    attachments_str,
                    email.message_id
                ])
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(output_path)
            self.logger.info(f"Excel file saved to {output_path}")
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {str(e)}")
    
    def append_to_existing_excel(self, emails: List[EmailData], excel_path: str) -> None:
        """Append data to an existing Excel file"""
        try:
            if os.path.exists(excel_path):
                workbook = openpyxl.load_workbook(excel_path)
                worksheet = workbook.active
            else:
                self.logger.warning(f"Excel file {excel_path} not found. Creating new file.")
                self.create_excel_from_emails(emails, excel_path)
                return
            
            # Find next empty row
            next_row = worksheet.max_row + 1
            
            # Add data rows
            for email in emails:
                body_preview = email.body[:100] + "..." if len(email.body) > 100 else email.body
                body_preview = body_preview.replace('\n', ' ').replace('\r', ' ')
                
                attachments_str = ", ".join(email.attachments) if email.attachments else "None"
                
                worksheet.append([
                    email.subject,
                    email.sender,
                    email.recipient,
                    email.date,
                    body_preview,
                    attachments_str,
                    email.message_id
                ])
            
            workbook.save(excel_path)
            self.logger.info(f"Data appended to {excel_path}")
            
        except Exception as e:
            self.logger.error(f"Error appending to Excel file: {str(e)}")


class AutomatedProcessor:
    """Main automation class that orchestrates the entire process"""
    
    def __init__(self, log_level: str = "INFO"):
        self.msg_processor = MSGProcessor(log_level)
        self.excel_writer = ExcelWriter()
        self.logger = logging.getLogger(__name__)
    
    def process_and_export(self, 
                          msg_folder: str, 
                          excel_output: str = "processed_emails.xlsx",
                          json_output: str = "processed_emails.json",
                          append_to_excel: bool = False) -> None:
        """Complete pipeline: process MSG files and export to Excel/JSON"""
        
        self.logger.info("Starting MSG processing pipeline...")
        
        # Extract data from MSG files
        emails = self.msg_processor.process_msg_folder(msg_folder)
        
        if not emails:
            self.logger.warning("No emails were processed successfully.")
            return
        
        self.logger.info(f"Successfully processed {len(emails)} emails")
        
        # Export to JSON
        self.msg_processor.convert_to_json(emails, json_output)
        
        # Export to Excel
        if append_to_excel and os.path.exists(excel_output):
            self.excel_writer.append_to_existing_excel(emails, excel_output)
        else:
            self.excel_writer.create_excel_from_emails(emails, excel_output)
        
        self.logger.info("Pipeline completed successfully!")
    
    def batch_process_folders(self, folder_paths: List[str], base_output_name: str = "batch_emails") -> None:
        """Process multiple folders of MSG files"""
        all_emails = []
        
        for folder_path in folder_paths:
            self.logger.info(f"Processing folder: {folder_path}")
            emails = self.msg_processor.process_msg_folder(folder_path)
            all_emails.extend(emails)
        
        if all_emails:
            excel_output = f"{base_output_name}.xlsx"
            json_output = f"{base_output_name}.json"
            
            self.msg_processor.convert_to_json(all_emails, json_output)
            self.excel_writer.create_excel_from_emails(all_emails, excel_output)
            
            self.logger.info(f"Batch processing completed. Processed {len(all_emails)} total emails.")


def main():
    """Example usage of the MSG processor"""
    
    # Configuration
    MSG_FOLDER = r"C:\Users\ADMIN\Downloads\W4_Aug_25"  # Update this path
    EXCEL_OUTPUT = "emails_export.xlsx"
    JSON_OUTPUT = "emails_export.json"
    
    # Create processor
    processor = AutomatedProcessor(log_level="INFO")
    
    # Process MSG files and export to Excel/JSON
    processor.process_and_export(
        msg_folder=MSG_FOLDER,
        excel_output=EXCEL_OUTPUT,
        json_output=JSON_OUTPUT,
        append_to_excel=False  # Set to True to append to existing Excel file
    )
    
    # Example of batch processing multiple folders
    # folders = [r"C:\path\to\folder1", r"C:\path\to\folder2"]
    # processor.batch_process_folders(folders, "batch_processed_emails")


if __name__ == "__main__":
    main()